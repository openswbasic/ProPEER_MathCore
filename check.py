# -*- coding: utf-8 -*-
"""
Created on Wed Jun 16 14:24:59 2021

@author: Park-HY-00

정/오답 체크 모듈

"""

from openpyxl import load_workbook

# 교재 답안 추출
answer_book = load_workbook("C:/Users/user/propeer/sample답.xlsx", data_only=True)
answer = answer_book["Sheet1"]
answer_name =answer.cell(row=1, column=1).value
book_Ans = []
for i in answer.rows:
    Ans = i[1].value
    book_Ans.append(Ans)

# 학생 작성 답안 추출
answer_std = load_workbook("C:/Users/user/propeer/sample학생답안.xlsx", data_only=True)
std1 = answer_std["이몽룡"]
std1_name = std1.cell(row=1, column=1).value
std1_Ans = []
for i in std1.rows:
    Ans = i[1].value
    std1_Ans.append(Ans)
std2 = answer_std["홍길동"]
std2_name = std2.cell(row=1, column=1).value
std2_Ans = []
for i in std2.rows:
    Ans = i[1].value
    std2_Ans.append(Ans)

# 정답 비교 및 엑셀 파일 기록
std1.correct = 0
std1.wrong = 0
for i in range(std1_Ans[1], std1_Ans[4]):
    if (std1_Ans[i] == book_Ans[i]):
        std1.corret += 1
        #std1.cell(row=3, column=i).value = 0
    else:
        std1.wrong += 1
        #std1.cell(row=3, column=i).value = 1

std2_correct = 0
std2_wrong = 0
for i in range(std2_Ans[1], std2_Ans[4]):
    for j in range(book_Ans[1], book_Ans[2]):
        if (std2_Ans[i]==book_Ans[j]):
            std2_correct += 1
            #std2.cell(row=3, column=i).value = 0
        else:
            std2_wrong += 1
            #std2.cell(row=3, column=i).value = 1


""" test code """
print("채점 교재 :", answer_name, ", 교재 답안 :", book_Ans)
print(std1_name,"의 채점 결과 :", std1_Ans)
print("정답 개수 :", std1.correct, "오답 개수 : ", std1.wrong)
print(std2_name,"의 채점 결과 :", std2_Ans)
print("정답 개수 :", std2_correct, "오답 개수 : ", std2_wrong)
